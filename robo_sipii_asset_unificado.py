"""
ROBÔ SIPII CAIXA — v11 (SIPII + CAIXA Asset JSON)
Estratégia de dados em três camadas:
  1. SIPII: rentabilidade, cota, PL, categorias e perfis mais atualizados
  2. URLs: dicionário estático + scraping dinâmico em páginas CAIXA
  3. CAIXA Asset JSON: CNPJ, SIICO, SIART, risco, benchmark, taxas, liquidez, captação e links oficiais
"""

from pathlib import Path
from datetime import datetime
import time, unicodedata, traceback, re, requests
from bs4 import BeautifulSoup
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException, InvalidSessionIdException, WebDriverException,
)

# ---------------------------------------------------------------------------
# Configurações Globais
# ---------------------------------------------------------------------------

URL_SIPII = "https://www.fundos.caixa.gov.br/sipii/pages/public/listar-fundos-internet.jsf"

PERFIS = [
    {"segmento": "PESSOA FÍSICA",   "sigla": "PF"},
    {"segmento": "PESSOA JURÍDICA", "sigla": "PJ"},
    {"segmento": "GOVERNO",         "sigla": "GOV"},
    {"segmento": "RPPS",            "sigla": "RPPS"},
    {"segmento": "TODOS",           "sigla": "TODOS"},
]

CATEGORIAS_FIXAS = [
    {"csv": "RENDA FIXA SIMPLES",           "texto_tela": "RENDA FIXA SIMPLES"},
    {"csv": "RENDA FIXA",                   "texto_tela": "RENDA FIXA"},
    {"csv": "RENDA FIXA REFERENCIADO",      "texto_tela": "RENDA FIXA REFERENCIADO"},
    {"csv": "RENDA FIXA CURTO PRAZO",       "texto_tela": "RENDA FIXA CURTO PRAZO"},
    {"csv": "MULTIMERCADO",                 "texto_tela": "MULTIMERCADO"},
    {"csv": "CAMBIAL",                      "texto_tela": "CAMBIAL"},
    {"csv": "ACOES",                        "texto_tela": "AÇÕES"},
    {"csv": "FUNDO DE INDICE",              "texto_tela": "FUNDO DE ÍNDICE"},
    {"csv": "FUNDOS MUTUOS DE PRIVATIZACAO","texto_tela": "FUNDOS MÚTUOS DE PRIVATIZAÇÃO"},
]

CAIXA_LISTING_PAGES = [
    "https://www.caixa.gov.br/fundos-investimento/renda-fixa/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/referenciados/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fundo-simples/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/curto-prazo/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/multimercado/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/cambiais/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fundos-de-indices/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/voce/renda-fixa/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/empresa/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/fmp-fgts/Paginas/default.aspx",
    "https://www.caixa.gov.br/fundos-investimento/rpps/Paginas/default.aspx",
]

# ---------------------------------------------------------------------------
# Dicionário estático — 116 URLs validadas manualmente
# ---------------------------------------------------------------------------
URL_ESTATICO = {
    # ── Validadas em 18/04/2026 (93) ────────────────────────────────────
    "CAIXA BRASIL INFLACAO ATIVA FIF RF CRED PRIV": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/brasil-IPCA",
    "CAIXA CAPITAL PROTEGIDO IBOVESPA CICLICO I FIC FIF MM": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-capital-protegido-ibovespa-ciclico-1-multimercado",
    "CAIXA ETF IBOVESPA FUNDO DE INDICE": "https://www.caixa.gov.br/fundos-investimento/fundos-de-indices/etf-ibovespa-fundo-de-indice/",
    "CAIXA EXPERT CLARITAS VALOR FIC FIF ACOES -": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/caixa-expert-claritas-valor-fic-acoes/Paginas/default.aspx",
    "CAIXA EXPERT GIANT ZARATHUSTRA FIC MULTIMERCADO -": "https://www.caixa.gov.br/fundos-investimento/multimercado/expert-giant-zarathustra-fic-multimercado",
    "CAIXA FI BRASIL IMA-B TP RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-b-titulos-rf-longo-prazo",
    "CAIXA FI BRASIL IRF-M 1 TP RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-irfm-1-titulos-publicos-rf",
    "CAIXA FI BRASIL IRF-M 1+ TP RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-irfm-1-titulos-publicos-rf-longo-prazo/",
    "CAIXA FIC ACOES EXPERT VERDE AM LONG BIAS -": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/CAIXA-Expert-Verde-Am-Long-Bias-FIC-Acoes/Paginas/default.aspx",
    "CAIXA FIC ACOES EXPERT VINCI VALOR DIVIDENDOS RPPS": "https://www.caixa.gov.br/fundos-investimento/rpps/caixa-fic-acoes-vinci-valor-dividendos-rpps",
    "CAIXA FIC ACOES EXPERT VINCI VALOR RPPS": "https://www.caixa.gov.br/fundos-investimento/rpps/caixa-fic-acoes-vinci-valor-rpps",
    "CAIXA FIC FIF ABSOLUTO PRE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-absoluto-pre-rf-longo-prazo",
    "CAIXA FIC FIF ACOES IBOVESPA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ibovespa",
    "CAIXA FIC FIF ACOES MULTIGESTOR": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fic-fia-multigestor/Paginas/default.aspx",
    "CAIXA FIC FIF ALOCACAO MACRO MM LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-alocacao-macro-multimercado-longo-prazo/",
    "CAIXA FIC FIF BETA RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-beta-ref-di-lp",
    "CAIXA FIC FIF BRASIL DISPONIBILIDADES SIMPLES RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/brasil-disponibilidades",
    "CAIXA FIC FIF BRASIL GESTAO ESTRATEGICA RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/caixa-fic-brasil-gestao-estrategica-rf",
    "CAIXA FIC FIF BRASIL RF REFER DI LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/referenciados/fi-brasil-ref-di-longo-prazo",
    "CAIXA FIC FIF CAPITAL IND PRECOS RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-capital-indice-de-precos-rf-longo-prazo",
    "CAIXA FIC FIF CLASSICO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-classico-rf-longo-prazo",
    "CAIXA FIC FIF DESENVOLVER RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-desenvolver-rf-longo-prazo/",
    "CAIXA FIC FIF E-FUNDO RF LP": "https://www.caixa.gov.br/fundos-investimento/e-fundos/renda-fixa-longo-prazo",
    "CAIXA FIC FIF EMPREENDER RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-empreender-rf-longo-prazo/",
    "CAIXA FIC FIF EQUILIBRIO MPE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/caixa-equilibrio-mpe-rf-lp/",
    "CAIXA FIC FIF ESTRATEGIA LIVRE MULTIMERCADO LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/estrategia-livre-mm-lp",
    "CAIXA FIC FIF ESTRATEGICO MULTIMERCADO LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-estrategico-multimercado-longo-prazo",
    "CAIXA FIC FIF EXECUTIVO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-executivo-rf-longo-prazo",
    "CAIXA FIC FIF EXPERTISE RF CRED PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-expertise-rf-credito-privado-longo-prazo",
    "CAIXA FIC FIF FACIL RF SIMPLES": "https://www.caixa.gov.br/fundos-investimento/fundo-simples/fic-facil-rf-simples",
    "CAIXA FIC FIF FOCO IND PRECOS RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-foco-indice-precos-rf-longo-prazo",
    "CAIXA FIC FIF FOF SMART MULTIESTRATEGIA MM": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-multigestor",
    "CAIXA FIC FIF GERACAO JOVEM CRED PRIV RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-geracao-jovem-rf-credito-privado-longo-prazo",
    "CAIXA FIC FIF GIRO IMEDIATO RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-giro-imediato-ref-di-longo-prazo",
    "CAIXA FIC FIF HEDGE MULTIMERCADO LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-fic-hedge-multimercado-lp/Paginas/default.aspx",
    "CAIXA FIC FIF IDEAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-ideal-rf-longo-prazo",
    "CAIXA FIC FIF INDEXA DOLAR CAMBIAL": "https://www.caixa.gov.br/fundos-investimento/cambiais/fic-cambial-dolar/Paginas/default.aspx",
    "CAIXA FIC FIF JUROS E MOEDAS MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-juros-moedas-multimercado-longo-prazo",
    "CAIXA FIC FIF JUROS E MOEDAS MM PLUS LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-fic-juros-e-moedas-multimercado-plus-lp/",
    "CAIXA FIC FIF NOVO BRASIL RF IMA-B LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-novo-brasil-ima-b-rf-longo-prazo",
    "CAIXA FIC FIF OAB RF CRED PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-oab-rf-credito-privado-longo-prazo",
    "CAIXA FIC FIF PATRIMONIO IND.DE PRECOS RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-patrimonio-indice-precos-rf-longo-prazo",
    "CAIXA FIC FIF PERSONAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-personal-rf-longo-prazo",
    "CAIXA FIC FIF PLENO REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-pleno-ref-di-longo-prazo",
    "CAIXA FIC FIF PRATICO RF CURTO PRAZO": "https://www.caixa.gov.br/fundos-investimento/curto-prazo/fic-pratico-curto-prazo",
    "CAIXA FIC FIF PREFERENCIAL REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-preferencial-ref-di-longo-prazo",
    "CAIXA FIC FIF RELACIONAMENTO PERSONAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-relacionamento-personal-rf-longo-prazo",
    "CAIXA FIC FIF RUBI RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-rubi-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF SELECAO RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-selecao-rf-longo-prazo",
    "CAIXA FIC FIF SIGMA RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-sigma-ref-di-longo-prazo",
    "CAIXA FIC FIF SOBERANO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-soberano-rf-longo-prazo",
    "CAIXA FIC FIF SUPREMO RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-supremo-rf-longo-prazo",
    "CAIXA FIC FIF TITULO PUBLICO MPE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/caixa-fic-tp-mpe-rf-lp/",
    "CAIXA FIC FIF TRANSFERENCIA VOLUNTARIA RF CP": "https://www.caixa.gov.br/fundos-investimento/curto-prazo/fic-transferencias-voluntarias-curto-prazo",
    "CAIXA FIC FIF TURQUESA CORPORATIVO RF CP": "https://www.caixa.gov.br/fundos-investimento/voce/renda-fixa/fic-turquesa-corporativo-curto-prazo",
    "CAIXA FIC FIFINVESTIDOR RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-investidor-rf-longo-prazo",
    "CAIXA FIC MULTIGESTOR GLOBAL EQUITIES IE": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-multigestor-global-equities-invest-ext/Paginas/default.aspx",
    "CAIXA FIC RELACIONAMENTO IDEAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-relacionamento-ideal-rf-longo-prazo",
    "CAIXA FIF ACOES BDR NIVEL I": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-bdr-nivel-1",
    "CAIXA FIF ACOES BRASIL ETF IBOVESPA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-brasil-etf-bovespa",
    "CAIXA FIF ACOES BRASIL IBOVESPA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-brasil-ibovespa/Paginas/default.aspx",
    "CAIXA FIF ACOES BRASIL IBX-50": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-brasil-ibx-50",
    "CAIXA FIF ACOES CONSTRUCAO CIVIL": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-construcao-civil",
    "CAIXA FIF ACOES CONSUMO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-consumo",
    "CAIXA FIF ACOES DIVIDENDOS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-dividendos",
    "CAIXA FIF ACOES IBOVESPA ATIVO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ibovespa-ativo",
    "CAIXA FIF ACOES IBRX ATIVO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ibrx-ativo",
    "CAIXA FIF ACOES INDEXA PIBB IBRX 50": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-pibb-sem-opcao-de-venda",
    "CAIXA FIF ACOES INDEXA SETOR FINANCEIRO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/caixa-fia-indexa-setor-financeiro/Paginas/default.aspx",
    "CAIXA FIF ACOES INFRAESTRUTURA": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-infraestrutura",
    "CAIXA FIF ACOES INSTITUCIONAL BDR NIVEL I": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/FIA-Institucional-BDR-nivel-I/Paginas/default.aspx",
    "CAIXA FIF ACOES PETROBRAS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-petrobras",
    "CAIXA FIF ACOES PETROBRAS PLUS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fia-caixa-petrobras-plus/Paginas/default.aspx",
    "CAIXA FIF ACOES PETROBRAS PRE-SAL": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-petrobras-pre-sal",
    "CAIXA FIF ACOES SMALL CAPS ATIVO": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-small-caps-ativo",
    "CAIXA FIF ACOES SUSTENTABILIDADE EMPRESARIAL ISE - IS": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-ise",
    "CAIXA FIF ACOES VALE DO RIO DOCE": "https://www.caixa.gov.br/fundos-investimento/fundos-de-acoes/fi-acoes-vale-rio-doce",
    "CAIXA FIF ALIANCA TP RF CURTO PRAZO -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-alianca-rf",
    "CAIXA FIF BRASIL IDKA IPCA 2A TIT PUB RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-idka-ipca-2a-rf-longo-prazo",
    "CAIXA FIF BRASIL IMA GERAL TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-geral-tp-rf-longo-prazo",
    "CAIXA FIF BRASIL IMA-B 5 TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-b-5-titulos-publicos-rf-longo-prazo",
    "CAIXA FIF BRASIL IMA-B 5+ TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-ima-b-5_mais-titulos-publicos-rf-lp/Paginas/default.aspx",
    "CAIXA FIF BRASIL IRF-M TP RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/brasil-irf-m-titulos-publicos-renda-fixa-longo-prazo",
    "CAIXA FIF BRASIL TITULOS PUBLICOS RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-brasil-titulos-publicos-rf-longo-prazo",
    "CAIXA FIF DIAMANTE CORP RF CRED PRIV LP -": "https://www.caixa.gov.br/fundos-investimento/empresa/renda-fixa/credito-privado/caixa-fi-diamante-corporativo-rf-cred-priv-lp",
    "CAIXA FIF E-SIMPLES RENDA FIXA LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-renda-fixa-simples-lp/",
    "CAIXA FIF INDEXA BOLSA AMERICANA MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-bolsa-americana-multimercado-lp",
    "CAIXA FIF INDEXA OURO MULTIMERCADO LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-ouro-multimercado-longo-prazo",
    "CAIXA FIF MULTIMERCADO RV 30 LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-multimercado-rv30-longo-prazo",
    "CAIXA FIF RS TITULOS PUBLICOS RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-rs-titulos-publicos-rf-longo-prazo",
    "CAIXA FIF SAUDE SUPLEMENTAR ANS II RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-saude-suplementar-ans-ii-rf-longo-prazo",
    "CAIXA FIF SAUDE SUPLEMENTAR ANS RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-saude-suplementar-anf-rf",
    "FIC FIF CAIXA EXPERT BTG PACTUAL X10 MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/btg-pactual-x10-multimercado-longo-prazo/",

    # ── Validadas em 30/04/2026 (23 novas) ──────────────────────────────
    "CAIXA FIC FIF OBJETIVO PRE RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-objetivo-pre-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF PERFORMANCE IMA-B RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-performance-ima-b-renda-fixa-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF PLUS QUALI RF CREDI PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-plus-qualificado-rf-credito-privado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF ESPECIAL RF LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-especial-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF FIDELIDADE PRIVATE RF LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-fidelidade-private-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF MAXI RENDA FIXA CRED PRIV LP -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fic-maxi-rf-credito-privado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF FIDELIDADE II RF CRED PRIV LP": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-fidelidade-ii-rf-credito-privado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC OMEGA REF DI RF LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-omega-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF MEGA RF REFERENC DI LONGO PRAZO -": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-mega-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF TOP PRIVATE RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-top-private-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF FOF SMART CRED PRIV MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-fof-smart-credito-privado-multimercado-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF INDEXA EURO MM LONGO PRAZO": "https://www.caixa.gov.br/fundos-investimento/multimercado/fi-euro-multimercado-longo-prazo/Paginas/default.aspx",
    "CAIXA CAPITAL PROTEGIDO CICLICO III FIC FIF MM LP -": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-capital-protegido-ciclico-iii-multimercado-lp/Paginas/default.aspx",
    "CAIXA FIC FIM CAPITAL PROTEGIDO CICLICO II LP -": "https://www.caixa.gov.br/fundos-investimento/multimercado/fic-capital-protegido-ciclico-ii-multimercado-lp/Paginas/default.aspx",
    "CAIXA FIC FIF EXPERT PIMCO INCOME IE MM LP": "https://www.caixa.gov.br/fundos-investimento/multimercado/caixa-expert-pimco-income-ie-fic-multimercado/Paginas/default.aspx",
    "CAIXA FIF EXTRAMERCADO COMUM IRFM-1 RF": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/fi-extramercado-comum-irfm-1-rf/Paginas/default.aspx",
    "CAIXA FIC FIF AMETISTA CORP RF SIMPLES": "https://www.caixa.gov.br/fundos-investimento/fundo-simples/fic-ametista-corporativo-rf-simples/Paginas/default.aspx",
    "CAIXA GIRO EMPRESAS FIC FIF RF REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-giro-empresas-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF CNI RF LP -": "https://www.caixa.gov.br/fundos-investimento/empresa/renda-fixa/fi-cni-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF SEBRAE RF LP": "https://www.caixa.gov.br/fundos-investimento/empresa/renda-fixa/fi-sebrae-rf-longo-prazo/Paginas/default.aspx",
    "CAIXA FIC FIF GIRO MPE REF DI LP": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-giro-mpe-ref-di-longo-prazo/Paginas/default.aspx",
    "CAIXA FIF TOPAZIO CORP RF REFERENC DI -": "https://www.caixa.gov.br/fundos-investimento/renda-fixa/referenciados/fic-topazio-rf-ref-di-lp/Paginas/default.aspx",
    "CAIXA FIC ESMERALDA CORP RF REF DI CRED PRIV LP -": "https://www.caixa.gov.br/fundos-investimento/referenciados/fic-esmeralda-corp-ref-di/Paginas/default.aspx",
}

DEBUG_COLUNAS = False

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
BASE_DIR = Path.cwd()
LOG_PATH = BASE_DIR / "execucao.log"

URL_ASSET_JSON = "https://www.caixa.gov.br/CAIXA-Asset/Documents/data/fundos.json"

DATA_DIR = BASE_DIR / "data"
HISTORICO_DIR = BASE_DIR / "historico"
HISTORICO_JSON_DIR = HISTORICO_DIR / "json"

DATA_DIR.mkdir(exist_ok=True)
HISTORICO_JSON_DIR.mkdir(parents=True, exist_ok=True)

def log(msg):
    linha = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(linha)
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(linha + "\n")

def normalizar(txt):
    txt = txt or ""
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return " ".join(txt.upper().split())

def chave_estatica(nome):
    n = normalizar(nome)
    n = re.sub(r'\s*RESP\s*LTDA.*$', '', n)
    n = re.sub(r'\s*-\s*RL$', '', n)
    n = re.sub(r'\s*\(\d+\).*$', '', n)
    return n.strip()

TEXTO_PARA_CSV = {normalizar(c["texto_tela"]): c["csv"] for c in CATEGORIAS_FIXAS}

# ---------------------------------------------------------------------------
# Camada 2 — Scraping dinâmico
# ---------------------------------------------------------------------------
STOPWORDS = {
    "CAIXA","FIC","FIF","FI","RF","LP","MM","IE","RL","IS",
    "RESP","LTDA","DE","DO","DA","DOS","DAS","E","A","O","EM",
    "FUNDO","FUNDOS","RENDA","FIXA","LONGO","PRAZO","CREDITO",
    "PRIVADO","TITULOS","PUBLICOS","TP","REFERENCIADO","SIMPLES",
}

def palavras_chave(texto):
    n = normalizar(texto)
    n = re.sub(r'\s*RESP\s*LTDA.*$', '', n)
    n = re.sub(r'\s*\(\d+\).*$', '', n)
    n = n.replace('-', ' ')
    tokens = set(n.split())
    return tokens - STOPWORDS - {t for t in tokens if len(t) <= 2}

def palavras_slug(url):
    try:
        path = url.rstrip('/').split('?')[0]
        parte = path.split('/')[-1]
        if parte.lower() in ('default.aspx', ''):
            parte = path.split('/')[-2]
        parte = re.sub(r'\.aspx$', '', parte).replace('-', ' ').replace('_', ' ')
        tokens = set(normalizar(parte).split())
        return tokens - STOPWORDS - {t for t in tokens if len(t) <= 2}
    except:
        return set()

HEADERS_HTTP = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/123.0.0.0 Safari/537.36",
    "Accept-Language": "pt-BR,pt;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.caixa.gov.br/",
}

def raspar_urls_caixa():
    registros, vistos = [], set()
    base = "https://www.caixa.gov.br"
    session = requests.Session()
    session.headers.update(HEADERS_HTTP)
    for page_url in CAIXA_LISTING_PAGES:
        try:
            resp = session.get(page_url, timeout=20)
            if resp.status_code != 200:
                log(f"  [CAIXA] {page_url.split('/')[-3]} → HTTP {resp.status_code}")
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            novos = 0
            for a in soup.find_all("a", href=True):
                href = a["href"].strip()
                texto = a.get_text(separator=" ", strip=True)
                if href.startswith("/"): href = base + href
                if not href.startswith("http"): continue
                if "/fundos-investimento/" not in href: continue
                if len(href.rstrip('/').split('/')) < 7: continue
                if href in vistos: continue
                vistos.add(href)
                registros.append({
                    "texto": normalizar(texto),
                    "url": href,
                    "palavras_texto": palavras_chave(texto),
                    "palavras_slug":  palavras_slug(href),
                })
                novos += 1
            log(f"  [CAIXA] {page_url.split('/')[-3]} → +{novos} links ({len(registros)} total)")
            time.sleep(1)
        except Exception as e:
            log(f"  [CAIXA] Erro em {page_url}: {e}")
    log(f"[CAIXA] Total URLs dinâmicas: {len(registros)}")
    return registros

def encontrar_url_dinamica(nome, registros):
    pf = palavras_chave(nome)
    if not pf: return ""
    melhor_url, melhor_score = "", 0.0
    for reg in registros:
        s = max(
            len(pf & reg["palavras_texto"]) / len(pf),
            len(pf & reg["palavras_slug"])  / len(pf),
        )
        if s > melhor_score:
            melhor_score, melhor_url = s, reg["url"]
    return melhor_url if melhor_score >= 0.70 else ""

def encontrar_url(nome, registros_dinamicos):
    chave = chave_estatica(nome)
    if chave in URL_ESTATICO:
        return URL_ESTATICO[chave]
    chave_curta = re.sub(r'\s+-\s*$', '', chave).strip()
    if chave_curta in URL_ESTATICO:
        return URL_ESTATICO[chave_curta]
    return encontrar_url_dinamica(nome, registros_dinamicos)


# ---------------------------------------------------------------------------
# CAIXA Asset JSON — download e enriquecimento
# ---------------------------------------------------------------------------
def baixar_json_asset():
    """
    Baixa o fundos.json da CAIXA Asset e salva:
      - data/fundos_asset.json
      - historico/json/fundos_asset_YYYYMMDD.json

    Se houver falha no download, tenta reaproveitar data/fundos_asset.json local.
    Se também não existir, retorna lista vazia para o robô seguir com SIPII puro.
    """
    import json

    log("ETAPA 4 — Baixando JSON da CAIXA Asset...")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/123.0.0.0 Safari/537.36",
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "pt-BR,pt;q=0.9",
        "Referer": "https://www.caixa.gov.br/caixa-asset/portfolio-fundos/Paginas/default.aspx",
    }

    caminho_atual = DATA_DIR / "fundos_asset.json"
    data_str = datetime.now().strftime("%Y%m%d")
    caminho_historico = HISTORICO_JSON_DIR / f"fundos_asset_{data_str}.json"

    try:
        resp = requests.get(URL_ASSET_JSON, headers=headers, timeout=60)
        resp.raise_for_status()
        dados = resp.json()

        with open(caminho_atual, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)

        with open(caminho_historico, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)

        log(f"  JSON Asset salvo: {caminho_atual}")
        log(f"  Histórico JSON salvo: {caminho_historico}")
        log(f"  Fundos no JSON Asset: {len(dados)}")
        return dados

    except Exception as e:
        log(f"  AVISO: não foi possível baixar o JSON Asset: {e}")

        if caminho_atual.exists():
            try:
                with open(caminho_atual, "r", encoding="utf-8") as f:
                    dados = json.load(f)
                log(f"  Usando JSON Asset local já existente: {caminho_atual} ({len(dados)} fundos)")
                return dados
            except Exception as e2:
                log(f"  AVISO: falha ao ler JSON Asset local: {e2}")

        log("  Seguindo sem enriquecimento da Asset.")
        return []


def limpar_cnpj(cnpj):
    if cnpj is None:
        return ""
    return "".join(filter(str.isdigit, str(cnpj)))


def normalizar_match(txt):
    """Normalização mais agressiva para cruzar nomes SIPII x Asset."""
    txt = normalizar(txt)

    remover = [
        "RESPONSABILIDADE LIMITADA",
        "RESP LTDA",
        "FUNDO DE INVESTIMENTO FINANCEIRO",
        "FUNDO DE INVESTIMENTO",
        "CLASSE DE",
        "EM COTAS DE",
        "FIC",
        "FIF",
        "FI",
        "RL",
        "LTDA",
    ]

    for termo in remover:
        txt = txt.replace(termo, " ")

    txt = re.sub(r"\(\d+\)", " ", txt)
    txt = re.sub(r"\s+-\s+", " ", txt)
    txt = re.sub(r"\s+", " ", txt)
    return txt.strip()


def sim_nao(valor):
    if valor is True:
        return "Sim"
    if valor is False:
        return "Não"
    return ""


def criar_indices_asset(dados_asset):
    """Cria índices por nome normalizado e CNPJ para localizar fundos do JSON Asset."""
    indice_nome = {}
    indice_cnpj = {}

    for item in dados_asset or []:
        nome = normalizar_match(item.get("no_fundo", ""))
        cnpj = limpar_cnpj(item.get("nu_cnpj", ""))

        if nome:
            indice_nome[nome] = item
        if cnpj:
            indice_cnpj[cnpj] = item

    return indice_nome, indice_cnpj


def encontrar_asset_por_nome(nome_sipii, indice_nome):
    """
    Tenta match exato por nome normalizado.
    Se não achar, tenta aproximação por interseção de palavras.
    """
    nome_norm = normalizar_match(nome_sipii)

    if nome_norm in indice_nome:
        return indice_nome[nome_norm], "nome_exato"

    palavras_sipii = set(nome_norm.split())
    melhor_item = None
    melhor_score = 0.0

    for nome_asset, item in indice_nome.items():
        palavras_asset = set(nome_asset.split())
        if not palavras_sipii:
            continue

        score = len(palavras_sipii & palavras_asset) / len(palavras_sipii)
        if score > melhor_score:
            melhor_score = score
            melhor_item = item

    if melhor_item and melhor_score >= 0.65:
        return melhor_item, f"nome_aproximado_{melhor_score:.2f}"

    return None, "sem_match"


def primeiro_link_valido(*links):
    for link in links:
        if link and str(link).strip() and str(link).strip().upper() != "INDISPONIVEL":
            return str(link).strip()
    return ""


def enriquecer_com_asset(df, dados_asset):
    """
    Acrescenta campos cadastrais/comerciais do JSON Asset ao df_final.
    Regra de governança:
      - SIPII continua sendo fonte principal para performance, cota, acumulados e PL.
      - JSON Asset complementa cadastro, risco, benchmark, taxa, liquidez, captação e links.
    """
    if df.empty:
        return df

    if not dados_asset:
        log("  JSON Asset vazio. Mantendo base apenas com dados SIPII.")
        df["Match Asset"] = "Não"
        return df

    indice_nome, indice_cnpj = criar_indices_asset(dados_asset)
    enriquecidos = []
    total_match = 0

    for _, row in df.iterrows():
        fundo = row.get("Fundo", "")
        extra, tipo_match = encontrar_asset_por_nome(fundo, indice_nome)

        if extra:
            total_match += 1

        novo = row.to_dict()

        link_pagina_asset = primeiro_link_valido(extra.get("de_link_pagina_fundo", "") if extra else "")
        link_lamina = primeiro_link_valido(extra.get("de_link_lamina", "") if extra else "")
        link_regulamento = primeiro_link_valido(extra.get("de_link_regulamento", "") if extra else "")
        link_termo = primeiro_link_valido(extra.get("de_link_termo_adesao", "") if extra else "")
        link_fato = primeiro_link_valido(extra.get("de_link_fato_relevante", "") if extra else "")
        link_info = primeiro_link_valido(extra.get("de_link_info_compl", "") if extra else "")
        link_sumario = primeiro_link_valido(extra.get("de_link_sumario", "") if extra else "")
        link_due = primeiro_link_valido(extra.get("de_due_diligence", "") if extra else "")
        link_raio_x = primeiro_link_valido(extra.get("de_link_raio_x", "") if extra else "")
        link_boletim = primeiro_link_valido(extra.get("de_link_boletim_comercial", "") if extra else "")

        novo.update({
            "Match Asset": "Sim" if extra else "Não",
            "Tipo Match Asset": tipo_match,

            "CNPJ": extra.get("nu_cnpj", "") if extra else "",
            "Código SIICO": extra.get("co_siico00", "") if extra else "",
            "Código SIICO Num": extra.get("co_siico", "") if extra else "",
            "Código SIART": extra.get("co_oferta_siart", "") if extra else "",
            "Oferta SIART": extra.get("co_oferta_siart", "") if extra else "",
            "Código MN": extra.get("co_mn", "") if extra else "",

            "Nome Asset": extra.get("no_fundo", "") if extra else "",
            "Razão Social": extra.get("no_razao_social", "") if extra else "",
            "Pesquisa Index Asset": extra.get("de_pesquisa_index", "") if extra else "",

            "Risco": extra.get("no_perfil_risco", "") if extra else "",
            "Benchmark": extra.get("no_benchmark", "") if extra else "",
            "Classificação CVM": extra.get("no_classificacao_cvm", "") if extra else "",
            "Estratégia": str(extra.get("no_estrategia", "")).strip() if extra else "",
            "Tributação": extra.get("no_classificacao_tributaria", "") if extra else "",
            "Classificação Investidor": extra.get("no_classificacao_investidor", "") if extra else "",

            "Taxa Adm Cliente (%)": extra.get("pc_taxa_adm_cliente", "") if extra else "",
            "Taxa Adm Máxima (%)": extra.get("pc_taxa_adm_maxima", "") if extra else "",

            "Aplicação Inicial (R$)": extra.get("vr_aplicacao_inicial", "") if extra else "",
            "Aplicação Adicional Mínima (R$)": extra.get("vr_aplicacao_adicional_minima", "") if extra else "",
            "Saldo Mínimo (R$)": extra.get("vr_saldo_minimo", "") if extra else "",
            "Resgate Mínimo (R$)": extra.get("vr_resgate_minimo", "") if extra else "",

            "Conversão Aplicação": extra.get("de_conversao_aplicacao", "") if extra else "",
            "Conversão Resgate": extra.get("de_conversao_resgate", "") if extra else "",
            "Pagamento Resgate": extra.get("de_pagamento_resgate", "") if extra else "",
            "Horário Resgate": extra.get("de_horario_resgate", "") if extra else "",
            "Horário Limite": extra.get("de_horario_limite", "") if extra else "",

            "Aberto Captação": sim_nao(extra.get("ic_aberto_captacao")) if extra else "",
            "Lançamento": sim_nao(extra.get("ic_lancamento")) if extra else "",
            "ASG": sim_nao(extra.get("ic_asg")) if extra else "",
            "Carência": sim_nao(extra.get("ic_carencia")) if extra else "",
            "Movimentação Automática": sim_nao(extra.get("ic_mov_automatica")) if extra else "",
            "Adiantamento Resgate": sim_nao(extra.get("ic_adiantamento_resgate")) if extra else "",
            "% Adiantamento Resgate": extra.get("pc_adiant_resgate", "") if extra else "",
            "Adiant. Manual/Automático": extra.get("de_adiant_manual_automatico", "") if extra else "",

            "Público Alvo Asset": extra.get("lista_publico_alvo", "") if extra else "",
            "Segmento Asset": extra.get("lista_segmento", "") if extra else "",
            "Destino": extra.get("de_destino", "") if extra else "",
            "Oferta IBC": extra.get("co_oferta_ibc", "") if extra else "",

            "Link Página Fundo": link_pagina_asset,
            "Link Lâmina": link_lamina,
            "Link Regulamento": link_regulamento,
            "Link Termo Adesão": link_termo,
            "Link Fato Relevante": link_fato,
            "Link Info Complementar": link_info,
            "Link Sumário": link_sumario,
            "Link Due Diligence": link_due,
            "Link Raio X": link_raio_x,
            "Link Boletim Comercial": link_boletim,

            "Observação Asset": extra.get("de_observacao_qs", "") if extra else "",
            "Data Cota Asset": extra.get("dt_cota_atual", "") if extra else "",
            "Data Início Asset": extra.get("dt_inicial", "") if extra else "",
        })

        # Se o robô não encontrou URL, usa a página oficial da Asset.
        if not novo.get("URL") and link_pagina_asset:
            novo["URL"] = link_pagina_asset

        enriquecidos.append(novo)

    log(f"  Asset matches encontrados: {total_match}/{len(df)}")
    log(f"  Asset sem match: {len(df) - total_match}/{len(df)}")

    return pd.DataFrame(enriquecidos)


def salvar_json_unificado(df):
    """Salva a base enriquecida em JSON para o HTML consumir diretamente."""
    import json

    data_str = datetime.now().strftime("%Y%m%d")
    caminho_atual = DATA_DIR / "fundos_unificado.json"
    caminho_hist = HISTORICO_JSON_DIR / f"fundos_unificado_{data_str}.json"

    registros = df.fillna("").to_dict(orient="records")

    with open(caminho_atual, "w", encoding="utf-8") as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)

    with open(caminho_hist, "w", encoding="utf-8") as f:
        json.dump(registros, f, ensure_ascii=False, indent=2)

    log(f"JSON unificado salvo: {caminho_atual}")
    log(f"Histórico JSON unificado salvo: {caminho_hist}")

# ---------------------------------------------------------------------------
# SIPII scraping
# ---------------------------------------------------------------------------
def configurar_driver(headless=True):
    opt = webdriver.ChromeOptions()
    if headless: opt.add_argument("--headless=new")
    for arg in ["--start-maximized","--window-size=1600,1200","--no-sandbox","--disable-dev-shm-usage","--disable-gpu"]:
        opt.add_argument(arg)
    opt.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/123.0.0.0 Safari/537.36")
    return webdriver.Chrome(options=opt)

def esperar_ajax(driver, timeout=20):
    try:
        WebDriverWait(driver, timeout).until(lambda d: d.execute_script("return (window.jQuery ? jQuery.active === 0 : true);"))
    except: time.sleep(2)

def encerrar_driver(driver):
    try:
        if driver: driver.quit()
    except: pass

def clicar_elemento(driver, el):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.5)
    try: el.click()
    except: driver.execute_script("arguments[0].click();", el)

def clicar_por_texto(driver, texto_alvo):
    alvo = normalizar(texto_alvo)
    for a in driver.find_elements(By.TAG_NAME, "a"):
        if a.is_displayed() and normalizar(a.text.strip()) == alvo:
            clicar_elemento(driver, a); return
    raise NoSuchElementException(f"Link não encontrado: {texto_alvo}")

def abrir_site_e_preparar(driver, sigla, segmento):
    log(f"[{sigla}] Abrindo SIPII...")
    driver.get(URL_SIPII); time.sleep(3)
    clicar_por_texto(driver, segmento); esperar_ajax(driver)
    consultar = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "btn-consultar")))
    clicar_elemento(driver, consultar); esperar_ajax(driver); time.sleep(3)

def descobrir_categorias(driver):
    cats = []
    for aba in driver.find_elements(By.CSS_SELECTOR, "ul.ui-tabs-nav li a"):
        texto = aba.text.strip()
        if texto:
            norm = normalizar(texto)
            cats.append({"texto_tela": texto, "csv": TEXTO_PARA_CSV.get(norm) or norm.replace(" ","_")[:31]})
    return cats

def localizar_tabela_ativa(driver):
    for p in driver.find_elements(By.CSS_SELECTOR, "div.ui-tabs-panel"):
        if p.is_displayed(): return p.find_element(By.CSS_SELECTOR, "table")
    raise Exception("Tabela ativa não encontrada")

def extrair_dados_tabela(driver, nome_csv, sigla):
    tabela = localizar_tabela_ativa(driver)
    linhas = tabela.find_elements(By.CSS_SELECTOR, "tbody tr")
    dados = []
    for i, tr in enumerate(linhas):
        tds = tr.find_elements(By.XPATH, "./td")
        if DEBUG_COLUNAS and i == 0:
            print("\n" + "="*60)
            for idx, td in enumerate(tds): print(f"  [{idx}] → '{td.text.strip()}'")
            print("="*60); return []
        if len(tds) >= 10:
            nome = tds[0].text.strip()
            if nome:
                dados.append({
                    "Categoria": nome_csv, "Fundo": nome,
                    "Fundo_norm": normalizar(nome),
                    "Data Inicio": tds[1].text.strip(),
                    "Cota (R$)": tds[3].text.strip(),
                    "Variacao Dia (%)": tds[4].text.strip(),
                    "Acum. Mes (%)": tds[5].text.strip(),
                    "Acum. Ano (%)": tds[6].text.strip(),
                    "Acum. 12M (%)": tds[7].text.strip(),
                    "PL (milhoes R$)": tds[8].text.strip(),
                    "Perfil": sigla,
                })
    return dados

def coletar_aba(driver, sigla, segmento, cat, dados):
    try:
        clicar_por_texto(driver, cat["texto_tela"]); esperar_ajax(driver); time.sleep(1.5)
        res = extrair_dados_tabela(driver, cat["csv"], sigla)
        dados.extend(res)
        log(f"  [{sigla}] {cat['csv']}: {len(res)} fundos.")
    except (InvalidSessionIdException, WebDriverException):
        log(f"  [{sigla}] Sessão perdida em '{cat['csv']}'. Reiniciando...")
        encerrar_driver(driver)
        driver = configurar_driver(headless=True)
        try:
            abrir_site_e_preparar(driver, sigla, segmento)
            clicar_por_texto(driver, cat["texto_tela"]); esperar_ajax(driver); time.sleep(1.5)
            res = extrair_dados_tabela(driver, cat["csv"], sigla)
            dados.extend(res); log(f"  [{sigla}] {cat['csv']} (recuperado): {len(res)} fundos.")
        except Exception as e2:
            log(f"  [{sigla}] Falha ao recuperar '{cat['csv']}': {e2}")
    except Exception as e:
        log(f"  [{sigla}] Erro em '{cat['csv']}': {e}")
    return driver

def processar_perfil(perfil, headless=True):
    sigla, segmento = perfil["sigla"], perfil["segmento"]
    driver, dados = None, []
    try:
        driver = configurar_driver(headless=headless)
        abrir_site_e_preparar(driver, sigla, segmento)
        for cat in descobrir_categorias(driver):
            driver = coletar_aba(driver, sigla, segmento, cat, dados)
    except Exception as e:
        log(f"[{sigla}] Erro geral: {e}"); traceback.print_exc()
    finally:
        encerrar_driver(driver)
    return dados

def consolidar(todos):
    if not todos: return pd.DataFrame()
    df = pd.DataFrame(todos)
    consolidado = []
    for (fn, cat), grupo in df.groupby(["Fundo_norm","Categoria"], sort=False):
        reg = grupo.iloc[0].to_dict()
        reg["Perfis"] = " | ".join(sorted(grupo["Perfil"].unique()))
        consolidado.append(reg)
    return pd.DataFrame(consolidado)

def salvar_excel(df, caminho):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fundos CAIXA"
        cols = [c for c in df.columns if c not in ("Fundo_norm","Perfil")]
        hf=PatternFill("solid",fgColor="0A1628"); hfont=Font(bold=True,color="C9A84C",size=10)
        ha=Alignment(horizontal="center",vertical="center",wrap_text=True)
        hb=Border(bottom=Side(style="medium",color="C9A84C"))
        ws.append(cols)
        for ci,cn in enumerate(cols,1):
            c=ws.cell(1,ci); c.fill=hf; c.font=hfont; c.alignment=ha; c.border=hb
        gf=Font(color="2EC27E",size=9); rf=Font(color="E05555",size=9); nf=Font(size=9)
        af=PatternFill("solid",fgColor="0F2040"); mf=PatternFill("solid",fgColor="0A1628")
        pc={"Variacao Dia (%)","Acum. Mes (%)","Acum. Ano (%)","Acum. 12M (%)"}
        for ri,row in enumerate(df[cols].itertuples(index=False),2):
            fill=af if ri%2==0 else mf
            for ci,(cn,val) in enumerate(zip(cols,row),1):
                vs=str(val) if val is not None else ""
                cell=ws.cell(ri,ci,vs); cell.fill=fill; cell.alignment=Alignment(vertical="center")
                if cn in pc and vs not in("-","—",""):
                    try:
                        n=float(vs.replace("%","").replace(" ","").replace(".","").replace(",","."))
                        cell.font=gf if n>0 else(rf if n<0 else nf)
                    except: cell.font=nf
                else: cell.font=nf
        widths={"Categoria":18,"Fundo":55,"Data Inicio":13,"Cota (R$)":16,
                "Variacao Dia (%)":14,"Acum. Mes (%)":13,"Acum. Ano (%)":13,
                "Acum. 12M (%)":13,"PL (milhoes R$)":18,"Perfis":30,"URL":60}
        for ci,cn in enumerate(cols,1):
            ws.column_dimensions[get_column_letter(ci)].width=widths.get(cn,15)
        ws.row_dimensions[1].height=30; ws.freeze_panes="A2"; wb.save(caminho)
        log(f"Excel salvo: {caminho}")
    except ImportError: log("AVISO: pip install openpyxl")
    except Exception as e: log(f"Erro Excel: {e}"); traceback.print_exc()

# ---------------------------------------------------------------------------
# Ponto de entrada
# ---------------------------------------------------------------------------
def executar():
    log("INICIANDO ROBÔ SIPII v11 - SIPII + ASSET JSON")

    if DEBUG_COLUNAS:
        driver = configurar_driver(headless=False)
        try:
            abrir_site_e_preparar(driver, "PF", "PESSOA FÍSICA")
            cats = descobrir_categorias(driver)
            if cats:
                clicar_por_texto(driver, cats[0]["texto_tela"])
                esperar_ajax(driver)
                time.sleep(1.5)
                extrair_dados_tabela(driver, cats[0]["csv"], "PF")
        finally:
            encerrar_driver(driver)
        return

    log(f"ETAPA 1 — Coletando URLs dinâmicas ({len(CAIXA_LISTING_PAGES)} páginas CAIXA)...")
    registros_dinamicos = raspar_urls_caixa()

    log("ETAPA 2 — Coletando dados do SIPII...")
    todos = []
    for p in PERFIS:
        todos.extend(processar_perfil(p, headless=True))

    df_final = consolidar(todos)
    if df_final.empty:
        log("ATENÇÃO: Nenhum dado coletado.")
        return

    log("ETAPA 3 — Associando URLs (estáticas + dinâmicas)...")
    df_final["URL"] = df_final["Fundo"].apply(lambda n: encontrar_url(n, registros_dinamicos))

    com_url = (df_final["URL"] != "").sum()
    sem_url = (df_final["URL"] == "").sum()
    log(f"  Dicionário estático : {len(URL_ESTATICO)} entradas")
    log(f"  URLs encontradas    : {com_url}/{len(df_final)}")
    log(f"  Sem URL             : {sem_url}/{len(df_final)}")

    if sem_url > 0:
        log("  Fundos SEM URL:")
        for n in df_final[df_final["URL"] == ""]["Fundo"].tolist():
            log(f"    - {n[:75]}")

    dados_asset = baixar_json_asset()
    df_final = enriquecer_com_asset(df_final, dados_asset)

    # Reordena colunas principais para o HTML e o Excel ficarem mais amigáveis.
    colunas_prioritarias = [
        "Categoria", "Fundo", "Fundo_norm", "Data Inicio", "Cota (R$)",
        "Variacao Dia (%)", "Acum. Mes (%)", "Acum. Ano (%)", "Acum. 12M (%)",
        "PL (milhoes R$)", "Perfil", "Perfis", "URL",
        "Match Asset", "Tipo Match Asset", "CNPJ", "Código SIICO", "Código SIICO Num",
        "Código SIART", "Oferta SIART", "Risco", "Benchmark", "Classificação CVM",
        "Estratégia", "Tributação", "Taxa Adm Cliente (%)", "Taxa Adm Máxima (%)",
        "Aplicação Inicial (R$)", "Aplicação Adicional Mínima (R$)", "Saldo Mínimo (R$)",
        "Resgate Mínimo (R$)", "Conversão Aplicação", "Conversão Resgate", "Pagamento Resgate",
        "Horário Limite", "Aberto Captação", "ASG", "Movimentação Automática",
        "Adiantamento Resgate", "Público Alvo Asset", "Segmento Asset",
        "Link Página Fundo", "Link Lâmina", "Link Regulamento", "Link Termo Adesão",
        "Link Fato Relevante", "Link Info Complementar", "Link Sumário",
        "Observação Asset", "Data Cota Asset"
    ]
    restantes = [c for c in df_final.columns if c not in colunas_prioritarias]
    colunas_final = [c for c in colunas_prioritarias if c in df_final.columns] + restantes
    df_final = df_final[colunas_final]

    data_str = datetime.now().strftime("%Y%m%d")

    caminho_csv = BASE_DIR / "dados_atuais.csv"
    df_final.to_csv(caminho_csv, index=False, encoding="utf-8-sig")
    log(f"CSV salvo: {caminho_csv}")

    caminho_csv_hist = BASE_DIR / f"sipii_caixa_{data_str}.csv"
    df_final.to_csv(caminho_csv_hist, index=False, encoding="utf-8-sig")
    log(f"Histórico CSV salvo: {caminho_csv_hist}")

    caminho_xlsx_hist = BASE_DIR / f"sipii_caixa_{data_str}.xlsx"
    salvar_excel(df_final, caminho_xlsx_hist)

    salvar_json_unificado(df_final)

    log(f"SUCESSO: {len(df_final)} fundos consolidados e enriquecidos salvos.")


if __name__ == "__main__":
    executar()
